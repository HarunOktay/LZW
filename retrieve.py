import os
import re
import openpyxl

def find_output_directories():
    """
    Finds all directories in the current working directory that match the pattern 'output_*'.
    Returns a list of directory paths.
    """
    current_dir = os.getcwd()
    dir_list = []
    pattern = re.compile(r'output_.*')

    for item in os.listdir(current_dir):
        if os.path.isdir(item) and pattern.match(item):
            dir_list.append(os.path.join(current_dir, item))

    return dir_list

def retrieve_excel_files(output_dirs):
    """
    From the list of output directories, retrieves the paths to the Excel files.
    Returns a list of tuples (excel_file_path, parameters_dict).
    """
    excel_files = []

    for dir_path in output_dirs:
        # Extract parameters from directory name
        dir_name = os.path.basename(dir_path)
        # Expected format: output_dict<max_dict_size>_code<code_bit_length>bit
        match = re.match(r'output_(dict\d+|nodictlimit)_code(\d+)bit', dir_name)
        if match:
            dict_size_str = match.group(1)
            code_length_str = match.group(2)

            # Convert parameters to appropriate types
            max_dict_size = None if dict_size_str == 'nodictlimit' else int(dict_size_str.replace('dict', ''))
            code_bit_length = int(code_length_str)

            # Look for Excel file in the directory
            for file_name in os.listdir(dir_path):
                if file_name.endswith('.xlsx'):
                    excel_file_path = os.path.join(dir_path, file_name)
                    parameters = {
                        'Directory': dir_name,
                        'Max Dictionary Size': max_dict_size if max_dict_size else 'No Limit',
                        'Code Bit Length': code_bit_length
                    }
                    excel_files.append((excel_file_path, parameters))
                    break  # Assuming only one Excel file per directory

    return excel_files

def aggregate_excel_files(excel_files):
    """
    Reads data from each Excel file and aggregates them.
    Returns a list of dictionaries containing the aggregated data.
    """
    aggregated_data = []

    for excel_file_path, parameters in excel_files:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        ws = wb.active

        # Skip the header row (assumed to be the first row)
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {
                'File Name': row[0],
                'Compressed File': row[1],
                'Original Size (bytes)': row[2],
                'Compressed Size (bytes)': row[3],
                'Compression Ratio': float(row[4]),
                'Compression Performance (%)': float(row[5]),
                'Max Dictionary Size': parameters['Max Dictionary Size'],
                'Code Bit Length': parameters['Code Bit Length']
            }
            aggregated_data.append(data)

    return aggregated_data

def save_aggregated_data(aggregated_data):
    """
    Saves the aggregated data to a new Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aggregated Results"

    # Write the headers
    headers = [
        'File Name',
        'Compressed File',
        'Original Size (bytes)',
        'Compressed Size (bytes)',
        'Compression Ratio',
        'Compression Performance (%)',
        'Max Dictionary Size',
        'Code Bit Length'
    ]
    ws.append(headers)

    # Write data rows
    for data in aggregated_data:
        ws.append([
            data['File Name'],
            data['Compressed File'],
            data['Original Size (bytes)'],
            data['Compressed Size (bytes)'],
            data['Compression Ratio'],
            data['Compression Performance (%)'],
            data['Max Dictionary Size'],
            data['Code Bit Length']
        ])

    # Save the aggregated Excel file
    output_file_name = 'Aggregated_Compression_Results.xlsx'
    wb.save(output_file_name)
    print(f"Aggregated results saved to '{output_file_name}'.")

if __name__ == "__main__":
    # Find all output directories
    output_dirs = find_output_directories()

    if not output_dirs:
        print("No output directories found.")
        exit(1)

    # Retrieve Excel files from the directories
    excel_files = retrieve_excel_files(output_dirs)

    if not excel_files:
        print("No Excel files found in the output directories.")
        exit(1)

    # Aggregate data from the Excel files
    aggregated_data = aggregate_excel_files(excel_files)

    if not aggregated_data:
        print("No data found in the Excel files.")
        exit(1)

    # Save aggregated data to a new Excel file
    save_aggregated_data(aggregated_data)